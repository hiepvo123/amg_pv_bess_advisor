from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

from utils import normalize_date, to_float


FIELD_MAP = {
    "Mực nước hồ lúc 0h00 (m)": "h0_m",
    "Mực nước hồ lúc 24h00 (m)": "h24_m",
    "Sai biệt MNH trong ngày (m)": "delta_h_reported_m",
    "Thể tích hồ lúc 24h00 (triệu m3)": "volume24_mcm",
    "Lưu lượng TB chạy máy (m3/s)": "q_turbine_m3s",
    "Lưu lượng TB vào hồ (m3/s)": "q_inflow_m3s",
    "Lưu lượng TB xả qua Đập (m3/s)": "q_spill_m3s",
    "Lưu lượng TB bốc hơi (m3/s)": "q_evap_m3s",
    "Lưu lượng xả dòng chảy tối thiểu (m3/s)": "q_minflow_m3s",
    "Lưu lượng xả dòng chảy tối thiểu (m3/s)": "q_minflow_m3s",
}

STANDARD_NUMERIC_COLUMNS = [
    "h0_m",
    "h24_m",
    "delta_h_reported_m",
    "volume24_mcm",
    "q_turbine_m3s",
    "q_inflow_m3s",
    "q_spill_m3s",
    "q_evap_m3s",
    "q_minflow_m3s",
]


def _select_data_sheet(workbook):
    """
    Pick the sheet that contains the reservoir table.
    Some months use Sheet1, some use Sheet2.
    """
    best_sheet = None
    best_score = -1

    for ws in workbook.worksheets:
        score = 0
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 20), values_only=True):
            row_text = " ".join(str(v) for v in row if v is not None)
            if "THÔNG SỐ THỦY VĂN" in row_text:
                score += 5
            if "Mực nước hồ" in row_text:
                score += 3
            if "Lưu lượng" in row_text:
                score += 2
        if score > best_score:
            best_score = score
            best_sheet = ws

    return best_sheet


def _parse_may_simple(ws, source_file):
    """
    Parse May file format:
    Date | H0 | evaporation
    """
    records = []

    for row in ws.iter_rows(values_only=True):
        date_value = normalize_date(row[0]) if len(row) >= 1 else None
        if date_value is None:
            continue

        records.append({
            "date": date_value,
            "source_file": source_file,
            "format_type": "may_simple",
            "h0_m": to_float(row[1]) if len(row) > 1 else None,
            "h24_m": None,
            "delta_h_reported_m": None,
            "volume24_mcm": None,
            "q_turbine_m3s": None,
            "q_inflow_m3s": None,
            "q_spill_m3s": None,
            "q_evap_m3s": to_float(row[2]) if len(row) > 2 else None,
            "q_minflow_m3s": None,
        })

    return records


def _parse_daily_blocks(ws, source_file):
    """
    Parse June/July block format.

    Layout example:
    A3 = date
    B5 = 'Mực nước hồ lúc 0h00 (m)', C5 = value
    B6 = 'Mực nước hồ lúc 24h00 (m)', C6 = value
    ... repeated for each day.
    """
    records = []
    rows = list(ws.iter_rows(values_only=True))
    i = 0

    while i < len(rows):
        row = rows[i]
        date_value = normalize_date(row[0]) if len(row) >= 1 else None

        if date_value is None:
            i += 1
            continue

        rec = {
            "date": date_value,
            "source_file": source_file,
            "format_type": "daily_blocks",
        }
        for col in STANDARD_NUMERIC_COLUMNS:
            rec[col] = None

        j = i + 1
        while j < len(rows):
            next_row = rows[j]

            # Stop when next daily block starts.
            next_date = normalize_date(next_row[0]) if len(next_row) >= 1 else None
            if next_date is not None:
                break

            label = str(next_row[1]).strip() if len(next_row) > 1 and next_row[1] is not None else ""
            value = next_row[2] if len(next_row) > 2 else None

            if label in FIELD_MAP:
                rec[FIELD_MAP[label]] = to_float(value)

            j += 1

        records.append(rec)
        i = j

    return records


def load_one_excel_file(file_path):
    """
    Read one Excel file and return standardized records.
    """
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    wb = load_workbook(file_path, data_only=True)
    ws = _select_data_sheet(wb)

    # May file has simple 3-column structure. June/July have daily blocks.
    if "5-2025" in file_path.name or "05-2025" in file_path.name:
        return _parse_may_simple(ws, file_path.name)

    return _parse_daily_blocks(ws, file_path.name)


def load_all_excel_files(input_files):
    """
    Load all monthly Excel files into one raw standardized dataframe.
    """
    all_records = []

    for file_path in input_files:
        records = load_one_excel_file(file_path)
        all_records.extend(records)

    df = pd.DataFrame(all_records)

    if df.empty:
        raise ValueError("No data was loaded. Check file paths and Excel layout.")

    return df
